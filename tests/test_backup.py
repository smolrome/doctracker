"""
tests/test_backup.py — Tests for services/backup.py.

Covers:
  - create_backup() returns dict with expected top-level keys
  - create_backup() meta includes version, created_at, counts
  - restore_backup() merge mode: inserts new docs, skips existing
  - restore_backup() replace mode: re-inserts even if doc already exists
  - _export_documents() returns all docs including deleted
  - _export_users() returns user records
  - _export_saved_offices() returns list from saved_offices.json
  - create_selective_backup() only includes requested items
  - create_excel_backup() returns valid Excel bytes
  - restore_backup() with invalid/missing doc skips gracefully
"""

import json
import os
import uuid
import pytest

os.environ.setdefault("SECRET_KEY", "h" * 32)
os.environ.setdefault("ADMIN_USERNAME", "testadmin")
os.environ.setdefault("ADMIN_PASSWORD", "AdminPass123!")
os.environ["DATABASE_URL"] = ""


@pytest.fixture(autouse=True)
def isolated(tmp_path, monkeypatch):
    """Each test gets a fresh tmp directory with pre-seeded JSON files."""
    # Seed required JSON files
    (tmp_path / "documents.json").write_text(json.dumps([]))
    (tmp_path / "users.json").write_text(json.dumps([]))
    (tmp_path / "saved_offices.json").write_text(json.dumps({}))
    os.environ["DATA_FILE"] = str(tmp_path / "documents.json")
    monkeypatch.chdir(tmp_path)


def _make_doc(name="Test Doc", status="Pending"):
    return {
        "id":       str(uuid.uuid4())[:8].upper(),
        "doc_id":   f"REF-2024-{uuid.uuid4().hex[:4].upper()}",
        "doc_name": name,
        "status":   status,
    }


# ── create_backup ─────────────────────────────────────────────────────────────

class TestCreateBackup:
    def test_returns_dict_with_required_keys(self):
        from services.backup import create_backup
        b = create_backup()
        assert isinstance(b, dict)
        for key in ("meta", "documents", "users", "routing_slips", "saved_offices"):
            assert key in b, f"Missing key: {key}"

    def test_meta_has_version_and_created_at(self):
        from services.backup import create_backup, BACKUP_VERSION
        b = create_backup()
        assert b["meta"]["version"] == BACKUP_VERSION
        assert "created_at" in b["meta"]

    def test_meta_counts_match_collections(self):
        from services.backup import create_backup
        from services.documents import insert_doc
        insert_doc(_make_doc("Doc A"))
        insert_doc(_make_doc("Doc B"))
        b = create_backup()
        assert b["meta"]["counts"]["documents"] == len(b["documents"])

    def test_counts_match_actual_collection_lengths(self):
        """meta.counts must equal the lengths of the actual collections."""
        from services.backup import create_backup
        b = create_backup()
        assert b["meta"]["counts"]["documents"] == len(b["documents"])
        assert b["meta"]["counts"]["users"]     == len(b["users"])
        assert b["meta"]["counts"]["routing_slips"] == len(b["routing_slips"])
        assert b["meta"]["counts"]["saved_offices"] == len(b["saved_offices"])

    def test_documents_list_contains_inserted_docs(self):
        from services.backup import create_backup
        from services.documents import insert_doc
        doc = _make_doc("Export Me")
        insert_doc(doc)
        b = create_backup()
        ids = [d["id"] for d in b["documents"]]
        assert doc["id"] in ids


# ── create_selective_backup ───────────────────────────────────────────────────

class TestSelectiveBackup:
    def test_only_requested_items_are_included(self):
        from services.backup import create_selective_backup
        b = create_selective_backup(["documents"])
        assert "documents" in b
        assert "users" not in b
        assert "routing_slips" not in b

    def test_selective_users_only(self):
        from services.backup import create_selective_backup
        b = create_selective_backup(["users"])
        assert "users" in b
        assert "documents" not in b

    def test_selective_offices_only(self):
        from services.backup import create_selective_backup
        b = create_selective_backup(["saved_offices"])
        assert "saved_offices" in b

    def test_selective_multi_items(self):
        from services.backup import create_selective_backup
        b = create_selective_backup(["documents", "users"])
        assert "documents" in b
        assert "users" in b
        assert "routing_slips" not in b

    def test_selective_date_filter_excludes_out_of_range(self):
        from services.backup import create_selective_backup
        from services.documents import insert_doc
        doc = _make_doc("Past Doc")
        doc["date_received"] = "2020-01-15"
        insert_doc(doc)
        b = create_selective_backup(["documents"], date_from="2025-01-01")
        # The 2020 doc should not be included
        ids = [d["id"] for d in b.get("documents", [])]
        assert doc["id"] not in ids


# ── restore_backup ────────────────────────────────────────────────────────────

class TestRestoreBackup:
    def test_restore_inserts_new_document(self):
        from services.backup import restore_backup
        from services.documents import get_doc
        doc = _make_doc("Restore Test")
        backup = {"documents": [doc], "users": [], "routing_slips": [], "saved_offices": []}
        summary = restore_backup(backup, mode="merge")
        assert summary["documents"] == 1
        fetched = get_doc(doc["id"])
        assert fetched is not None
        assert fetched["doc_name"] == "Restore Test"

    def test_merge_mode_skips_existing_document(self):
        from services.backup import restore_backup
        from services.documents import insert_doc
        doc = _make_doc("Already Exists")
        insert_doc(doc)
        backup = {"documents": [doc], "users": [], "routing_slips": [], "saved_offices": []}
        summary = restore_backup(backup, mode="merge")
        assert summary["skipped"] >= 1
        assert summary["documents"] == 0

    def test_replace_mode_inserts_over_existing(self):
        """Replace mode still inserts if get_doc() finds it (batch_save_docs upserts)."""
        from services.backup import restore_backup
        from services.documents import insert_doc
        doc = _make_doc("Replace Test")
        insert_doc(doc)
        backup = {"documents": [doc], "users": [], "routing_slips": [], "saved_offices": []}
        summary = restore_backup(backup, mode="replace")
        # In replace mode no skip; valid_docs = [doc] → batch_save called
        assert summary["skipped"] == 0

    def test_restore_invalid_doc_is_skipped(self):
        from services.backup import restore_backup
        invalid = {"no_id_here": "bad_record"}
        backup = {"documents": [invalid], "users": [], "routing_slips": [], "saved_offices": []}
        summary = restore_backup(backup, mode="merge")
        assert summary["skipped"] == 1

    def test_restore_empty_backup(self):
        from services.backup import restore_backup
        backup = {"documents": [], "users": [], "routing_slips": [], "saved_offices": []}
        summary = restore_backup(backup, mode="merge")
        assert summary["documents"] == 0

    def test_restore_summary_has_expected_keys(self):
        from services.backup import restore_backup
        summary = restore_backup({}, mode="merge")
        for key in ("mode", "documents", "users", "routing_slips", "saved_offices", "skipped", "errors"):
            assert key in summary


# ── _export helpers ───────────────────────────────────────────────────────────

class TestExportHelpers:
    def test_export_documents_includes_deleted(self):
        from services.backup import _export_documents
        from services.documents import insert_doc, delete_doc
        doc = _make_doc("Soft Deleted")
        insert_doc(doc)
        delete_doc(doc["id"])
        all_docs = _export_documents()
        ids = [d["id"] for d in all_docs]
        assert doc["id"] in ids

    def test_export_saved_offices_from_json(self):
        import json as _json
        # Seed the offices file
        data = {"main-office": {"office_name": "Main Office", "created_by": "admin"}}
        with open("saved_offices.json", "w") as f:
            _json.dump(data, f)
        from services.backup import _export_saved_offices
        offices = _export_saved_offices()
        assert isinstance(offices, list)
        assert any(o["office_slug"] == "main-office" for o in offices)

    def test_export_users_returns_list(self):
        from services.backup import _export_users
        users = _export_users()
        assert isinstance(users, list)


# ── Excel export ──────────────────────────────────────────────────────────────

class TestExcelBackup:
    def test_create_excel_backup_returns_bytes(self):
        pytest.importorskip("openpyxl")
        from services.backup import create_excel_backup
        result = create_excel_backup()
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_excel_bytes_are_valid_xlsx(self):
        pytest.importorskip("openpyxl")
        from services.backup import create_excel_backup
        from openpyxl import load_workbook
        from io import BytesIO
        data = create_excel_backup()
        wb = load_workbook(BytesIO(data))
        # Must have at least the Summary sheet
        assert "Summary" in wb.sheetnames
        assert "Documents" in wb.sheetnames

    def test_excel_with_one_document_has_data_row(self):
        pytest.importorskip("openpyxl")
        from services.backup import create_excel_backup
        from services.documents import insert_doc
        from openpyxl import load_workbook
        from io import BytesIO
        doc = _make_doc("Excel Row Doc")
        insert_doc(doc)
        data = create_excel_backup()
        wb = load_workbook(BytesIO(data))
        ws = wb["Documents"]
        # Row 3 = header, row 4 = first data row
        val = ws.cell(row=4, column=7).value  # doc_name column
        assert val == "Excel Row Doc"
