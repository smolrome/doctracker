"""
services/backup.py — Full system backup and restore.

Backup exports everything into a single JSON file the admin downloads.
Restore uploads that file and merges/replaces data back into the DB.

Backup includes:
  - All documents (including deleted/trashed)
  - All users (without password hashes for security)
  - All routing slips
  - All saved offices
  - All office traffic logs
  - Backup metadata (timestamp, version, doc count)
"""

import json
from datetime import datetime
from io import BytesIO

from services.database import USE_DB, get_conn
from services.documents import load_docs, save_doc, insert_doc, get_doc, now_str

BACKUP_VERSION = "2"


# ── Excel export ──────────────────────────────────────────────────────────────

def create_excel_backup() -> bytes:
    """Export all data to a formatted multi-sheet Excel workbook. Returns bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, GradientFill)
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── shared styles ──
    NAVY   = "0A2540"
    TEAL   = "0E7490"
    GOLD   = "C8922A"
    LIGHT  = "EFF6FF"
    WHITE  = "FFFFFF"
    GRAY   = "F8FAFC"

    hdr_font    = Font(bold=True, color=WHITE, name="Arial", size=10)
    hdr_fill    = PatternFill("solid", fgColor=NAVY)
    title_font  = Font(bold=True, color=NAVY, name="Arial", size=13)
    sub_font    = Font(bold=True, color=TEAL, name="Arial", size=10)
    cell_font   = Font(name="Arial", size=9)
    alt_fill    = PatternFill("solid", fgColor=GRAY)
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left        = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin_border = Border(
        bottom=Side(style="thin", color="CBD5E1"),
    )

    def style_header_row(ws, row_num, col_count):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.font  = hdr_font
            cell.fill  = hdr_fill
            cell.alignment = center
            cell.border    = thin_border

    def style_data_row(ws, row_num, col_count, alternate=False):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.font      = cell_font
            cell.alignment = left
            cell.border    = thin_border
            if alternate:
                cell.fill = alt_fill

    def set_col_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def add_title(ws, title, subtitle=""):
        ws.row_dimensions[1].height = 32
        ws["A1"] = title
        ws["A1"].font      = title_font
        ws["A1"].alignment = left
        if subtitle:
            ws["A2"] = subtitle
            ws["A2"].font      = Font(name="Arial", size=9, color="64748B", italic=True)
            ws["A2"].alignment = left
        ws["A1"].fill = PatternFill("solid", fgColor=LIGHT)

    # ═══════════════════════════════════════════════════════
    #  SHEET 1 — Summary
    # ═══════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    docs    = _export_documents()
    users   = _export_users()
    slips   = _export_routing_slips()
    offices = _export_saved_offices()
    now     = datetime.now().strftime("%B %d, %Y %I:%M %p")

    ws["A1"] = "DocTracker — DepEd Leyte Division"
    ws["A1"].font = Font(bold=True, name="Arial", size=16, color=NAVY)
    ws["A2"] = f"Data Export  ·  {now}"
    ws["A2"].font = Font(name="Arial", size=10, color="64748B", italic=True)
    ws.row_dimensions[1].height = 30

    summary_data = [
        ("", ""),
        ("Category", "Count"),
        ("Documents (total)", len(docs)),
        ("  Active",          sum(1 for d in docs if not d.get("deleted"))),
        ("  Deleted / Trash", sum(1 for d in docs if d.get("deleted"))),
        ("", ""),
        ("Document Status", ""),
        ("  Pending",    sum(1 for d in docs if d.get("status") == "Pending")),
        ("  Received",   sum(1 for d in docs if d.get("status") == "Received")),
        ("  Routed", sum(1 for d in docs if d.get("status") == "Routed")),
        ("  In Review",  sum(1 for d in docs if d.get("status") == "In Review")),
        ("  Released",   sum(1 for d in docs if d.get("status") == "Released")),
        ("  On Hold",    sum(1 for d in docs if d.get("status") == "On Hold")),
        ("  Archived",   sum(1 for d in docs if d.get("status") == "Archived")),
        ("", ""),
        ("Users",          len(users)),
        ("Routing Slips",  len(slips)),
        ("Saved Offices",  len(offices)),
    ]

    for i, (label, val) in enumerate(summary_data, start=4):
        ws.cell(row=i, column=1, value=label).font = Font(name="Arial", size=10,
            bold=(label in ("Category", "Document Status", "Documents (total)",
                            "Users", "Routing Slips", "Saved Offices")),
            color=NAVY)
        cell = ws.cell(row=i, column=2, value=val)
        cell.font = Font(name="Arial", size=10, bold=isinstance(val, int) and val > 0)
        cell.alignment = center
        if label == "Category":
            ws.cell(row=i, column=1).fill = PatternFill("solid", fgColor=NAVY)
            ws.cell(row=i, column=1).font = hdr_font
            ws.cell(row=i, column=2).fill = PatternFill("solid", fgColor=NAVY)
            ws.cell(row=i, column=2).font = hdr_font

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14

    # ═══════════════════════════════════════════════════════
    #  SHEET 2 — Documents
    # ═══════════════════════════════════════════════════════
    wd = wb.create_sheet("Documents")
    wd.sheet_view.showGridLines = False
    add_title(wd, "All Documents", f"Exported {now}")

    doc_headers = ["#", "Date Received", "Time", "Received By",
                   "Unit / Office / School / District", "Source / Sender",
                   "Document / Content Particulars",
                   "Referred To", "Forwarded To", "Date & Release Time",
                   "Status", "Category", "Reference No.", "Notes"]
    doc_widths  = [5, 14, 12, 22, 30, 26, 48, 22, 22, 18, 14, 16, 14, 30]

    for c, h in enumerate(doc_headers, 1):
        wd.cell(row=3, column=c, value=h)
    style_header_row(wd, 3, len(doc_headers))
    set_col_widths(wd, doc_widths)
    wd.freeze_panes = "A4"

    active_docs = [d for d in docs if not d.get("deleted")]
    for r, doc in enumerate(active_docs, start=4):
        alt = (r % 2 == 0)
        row = [
            r - 3,
            doc.get("date_received", ""),
            (doc.get("created_at", "") or "")[11:16],
            doc.get("received_by", ""),
            doc.get("sender_org", ""),
            doc.get("sender_name", ""),
            doc.get("doc_name", ""),
            doc.get("referred_to", ""),
            doc.get("forwarded_to", ""),
            doc.get("date_released", ""),
            doc.get("status", ""),
            doc.get("category", ""),
            doc.get("doc_id", ""),
            doc.get("notes", ""),
        ]
        for c, val in enumerate(row, 1):
            wd.cell(row=r, column=c, value=val)
        style_data_row(wd, r, len(doc_headers), alt)
        # Color-code status
        status_colors = {
            "Released":   ("D1FAE5", "065F46"),
            "Routed": ("DBEAFE", "1D4ED8"),
            "Received":   ("E0F2FE", "0369A1"),
            "Pending":    ("FEF3C7", "92400E"),
            "On Hold":    ("FEE2E2", "991B1B"),
            "In Review":  ("F3E8FF", "6B21A8"),
            "Archived":   ("F1F5F9", "475569"),
        }
        status = doc.get("status", "")
        if status in status_colors:
            bg, fg = status_colors[status]
            wd.cell(row=r, column=11).fill = PatternFill("solid", fgColor=bg)
            wd.cell(row=r, column=11).font = Font(name="Arial", size=9,
                                                  bold=True, color=fg)
    wd.auto_filter.ref = f"A3:{get_column_letter(len(doc_headers))}3"

    # ═══════════════════════════════════════════════════════
    #  SHEET 3 — Trash (deleted docs)
    # ═══════════════════════════════════════════════════════
    deleted = [d for d in docs if d.get("deleted")]
    if deleted:
        wt = wb.create_sheet("Trash")
        wt.sheet_view.showGridLines = False
        add_title(wt, "Deleted Documents (Trash)", f"Exported {now}")
        trash_headers = ["#", "ID", "Reference No.", "Document / Content",
                         "Status", "Deleted By", "Deleted At"]
        for c, h in enumerate(trash_headers, 1):
            wt.cell(row=3, column=c, value=h)
        style_header_row(wt, 3, len(trash_headers))
        set_col_widths(wt, [5, 10, 16, 44, 14, 20, 18])
        for r, doc in enumerate(deleted, start=4):
            row = [r - 3, doc.get("id",""), doc.get("doc_id",""),
                   doc.get("doc_name",""), doc.get("status",""),
                   doc.get("deleted_by",""), (doc.get("deleted_at","") or "")[:16]]
            for c, val in enumerate(row, 1):
                wt.cell(row=r, column=c, value=val).font = cell_font

    # ═══════════════════════════════════════════════════════
    #  SHEET 4 — Routing Slips
    # ═══════════════════════════════════════════════════════
    wr = wb.create_sheet("Routing Slips")
    wr.sheet_view.showGridLines = False
    add_title(wr, "Routing Slips", f"Exported {now}")

    slip_headers = ["#", "Slip No.", "Date", "Destination",
                    "Prepared By", "Time From", "Time To",
                    "No. of Docs", "Notes"]
    slip_widths  = [5, 18, 14, 30, 24, 12, 12, 10, 36]
    for c, h in enumerate(slip_headers, 1):
        wr.cell(row=3, column=c, value=h)
    style_header_row(wr, 3, len(slip_headers))
    set_col_widths(wr, slip_widths)
    wr.freeze_panes = "A4"

    for r, slip in enumerate(slips, start=4):
        alt = (r % 2 == 0)
        row = [
            r - 3,
            slip.get("slip_no", ""),
            slip.get("slip_date", "") or (slip.get("created_at","") or "")[:10],
            slip.get("destination", ""),
            slip.get("prepared_by", ""),
            slip.get("time_from", ""),
            slip.get("time_to", ""),
            len(slip.get("doc_ids", [])),
            slip.get("notes", ""),
        ]
        for c, val in enumerate(row, 1):
            wr.cell(row=r, column=c, value=val)
        style_data_row(wr, r, len(slip_headers), alt)

    # ═══════════════════════════════════════════════════════
    #  SHEET 5 — Users
    # ═══════════════════════════════════════════════════════
    wu = wb.create_sheet("Users")
    wu.sheet_view.showGridLines = False
    add_title(wu, "User Accounts", f"Exported {now}")

    user_headers = ["#", "Username", "Full Name", "Role", "Office / Unit",
                    "Active", "Last Login", "Created At"]
    for c, h in enumerate(user_headers, 1):
        wu.cell(row=3, column=c, value=h)
    style_header_row(wu, 3, len(user_headers))
    set_col_widths(wu, [5, 20, 28, 12, 30, 8, 20, 20])
    wu.freeze_panes = "A4"

    for r, u in enumerate(users, start=4):
        alt = (r % 2 == 0)
        row = [
            r - 3,
            u.get("username", ""),
            u.get("full_name", ""),
            u.get("role", ""),
            u.get("office", ""),
            "Yes" if u.get("active", True) else "No",
            str(u.get("last_login", "") or "")[:16],
            str(u.get("created_at", "") or "")[:16],
        ]
        for c, val in enumerate(row, 1):
            wu.cell(row=r, column=c, value=val)
        style_data_row(wu, r, len(user_headers), alt)

    # ═══════════════════════════════════════════════════════
    #  SHEET 6 — Offices
    # ═══════════════════════════════════════════════════════
    wo = wb.create_sheet("Saved Offices")
    wo.sheet_view.showGridLines = False
    add_title(wo, "Saved Offices", f"Exported {now}")

    office_headers = ["#", "Office Slug", "Office Name", "Created By", "Created At"]
    for c, h in enumerate(office_headers, 1):
        wo.cell(row=3, column=c, value=h)
    style_header_row(wo, 3, len(office_headers))
    set_col_widths(wo, [5, 28, 32, 20, 18])

    for r, o in enumerate(offices, start=4):
        alt = (r % 2 == 0)
        row = [r - 3, o.get("office_slug",""), o.get("office_name",""),
               o.get("created_by",""), str(o.get("created_at","") or "")[:16]]
        for c, val in enumerate(row, 1):
            wo.cell(row=r, column=c, value=val)
        style_data_row(wo, r, len(office_headers), alt)

    # Save to bytes
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── Export ────────────────────────────────────────────────────────────────────

def create_backup() -> dict:
    """Collect all data from the database into a single dict."""
    backup = {
        "meta": {
            "version":    BACKUP_VERSION,
            "created_at": datetime.now().isoformat(),
            "app":        "DocTracker - DepEd Leyte Division",
        },
        "documents":     _export_documents(),
        "users":         _export_users(),
        "routing_slips": _export_routing_slips(),
        "saved_offices": _export_saved_offices(),
        "office_traffic": _export_office_traffic(),
    }
    backup["meta"]["counts"] = {
        "documents":     len(backup["documents"]),
        "users":         len(backup["users"]),
        "routing_slips": len(backup["routing_slips"]),
        "saved_offices": len(backup["saved_offices"]),
    }
    return backup


def create_selective_backup(export_items: list) -> dict:
    """Collect selected data from the database into a single dict."""
    backup = {
        "meta": {
            "version":    BACKUP_VERSION,
            "created_at": datetime.now().isoformat(),
            "app":        "DocTracker - DepEd Leyte Division",
            "export_type": "selective",
            "items":      export_items,
        },
    }
    
    counts = {}
    if "documents" in export_items:
        backup["documents"] = _export_documents()
        counts["documents"] = len(backup["documents"])
    if "users" in export_items:
        backup["users"] = _export_users()
        counts["users"] = len(backup["users"])
    if "routing_slips" in export_items:
        backup["routing_slips"] = _export_routing_slips()
        counts["routing_slips"] = len(backup["routing_slips"])
    if "saved_offices" in export_items:
        backup["saved_offices"] = _export_saved_offices()
        counts["saved_offices"] = len(backup["saved_offices"])
    if "office_traffic" in export_items:
        backup["office_traffic"] = _export_office_traffic()
        counts["office_traffic"] = len(backup["office_traffic"])
    
    backup["meta"]["counts"] = counts
    return backup


def create_selective_excel_backup(export_items: list) -> bytes:
    """Export selected data to a formatted multi-sheet Excel workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    
    NAVY   = "0A2540"
    WHITE  = "FFFFFF"
    GRAY   = "F8FAFC"
    LIGHT  = "EFF6FF"
    
    hdr_font    = Font(bold=True, color=WHITE, name="Arial", size=10)
    hdr_fill    = PatternFill("solid", fgColor=NAVY)
    title_font  = Font(bold=True, color=NAVY, name="Arial", size=13)
    cell_font   = Font(name="Arial", size=9)
    alt_fill    = PatternFill("solid", fgColor=GRAY)
    left        = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(bottom=Side(style="thin", color="CBD5E1"))
    
    def style_header_row(ws, row_num, col_count):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.font  = hdr_font
            cell.fill  = hdr_fill
            cell.alignment = center
            cell.border    = thin_border

    def style_data_row(ws, row_num, col_count, alternate=False):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.font      = cell_font
            cell.alignment = left
            cell.border    = thin_border
            if alternate:
                cell.fill = alt_fill

    def set_col_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def add_title(ws, title, subtitle=""):
        ws.row_dimensions[1].height = 32
        ws["A1"] = title
        ws["A1"].font      = title_font
        ws["A1"].alignment = left
        if subtitle:
            ws["A2"] = subtitle
            ws["A2"].font      = Font(name="Arial", size=9, color="64748B", italic=True)
            ws["A2"].alignment = left
        ws["A1"].fill = PatternFill("solid", fgColor=LIGHT)

    now = datetime.now().strftime("%B %d, %Y %I:%M %p")
    
    # Documents sheet
    if "documents" in export_items:
        docs = _export_documents()
        wd = wb.create_sheet("Documents")
        wb.active = wd
        wd.sheet_view.showGridLines = False
        add_title(wd, "Documents", f"Exported {now}")
        
        doc_headers = ["#", "Date Received", "Received By", "Unit / Office",
                       "Source / Sender", "Document / Content",
                       "Referred To", "Status", "Category", "Reference No."]
        doc_widths  = [5, 14, 22, 30, 26, 48, 22, 14, 16, 14]
        
        for c, h in enumerate(doc_headers, 1):
            wd.cell(row=3, column=c, value=h)
        style_header_row(wd, 3, len(doc_headers))
        set_col_widths(wd, doc_widths)
        wd.freeze_panes = "A4"
        
        active_docs = [d for d in docs if not d.get("deleted")]
        for r, doc in enumerate(active_docs, start=4):
            alt = (r % 2 == 0)
            row = [r - 3, doc.get("date_received", ""), doc.get("received_by", ""),
                   doc.get("sender_org", ""), doc.get("sender_name", ""),
                   doc.get("doc_name", ""), doc.get("referred_to", ""),
                   doc.get("status", ""), doc.get("category", ""), doc.get("doc_id", "")]
            for c, val in enumerate(row, 1):
                wd.cell(row=r, column=c, value=val)
            style_data_row(wd, r, len(doc_headers), alt)
    
    # Trash sheet
    if "documents" in export_items:
        docs = _export_documents()
        deleted = [d for d in docs if d.get("deleted")]
        if deleted:
            wt = wb.create_sheet("Trash")
            wt.sheet_view.showGridLines = False
            add_title(wt, "Deleted Documents (Trash)", f"Exported {now}")
            trash_headers = ["#", "ID", "Reference No.", "Document", "Status", "Deleted By", "Deleted At"]
            for c, h in enumerate(trash_headers, 1):
                wt.cell(row=3, column=c, value=h)
            style_header_row(wt, 3, len(trash_headers))
            for r, doc in enumerate(deleted, start=4):
                row = [r - 3, doc.get("id",""), doc.get("doc_id",""), doc.get("doc_name",""),
                       doc.get("status",""), doc.get("deleted_by",""), (doc.get("deleted_at","") or "")[:16]]
                for c, val in enumerate(row, 1):
                    wt.cell(row=r, column=c, value=val).font = cell_font
    
    # Routing Slips sheet
    if "routing_slips" in export_items:
        slips = _export_routing_slips()
        wr = wb.create_sheet("Routing Slips")
        wr.sheet_view.showGridLines = False
        add_title(wr, "Routing Slips", f"Exported {now}")
        
        slip_headers = ["#", "Slip No.", "Date", "Destination", "Prepared By", "No. of Docs", "Notes"]
        for c, h in enumerate(slip_headers, 1):
            wr.cell(row=3, column=c, value=h)
        style_header_row(wr, 3, len(slip_headers))
        wr.freeze_panes = "A4"
        
        for r, slip in enumerate(slips, start=4):
            alt = (r % 2 == 0)
            row = [r - 3, slip.get("slip_no", ""), slip.get("slip_date", "") or (slip.get("created_at","") or "")[:10],
                   slip.get("destination", ""), slip.get("prepared_by", ""), len(slip.get("doc_ids", [])), slip.get("notes", "")]
            for c, val in enumerate(row, 1):
                wr.cell(row=r, column=c, value=val)
            style_data_row(wr, r, len(slip_headers), alt)
    
    # Users sheet
    if "users" in export_items:
        users = _export_users()
        wu = wb.create_sheet("Users")
        wu.sheet_view.showGridLines = False
        add_title(wu, "User Accounts", f"Exported {now}")
        
        user_headers = ["#", "Username", "Full Name", "Role", "Office", "Active", "Last Login"]
        for c, h in enumerate(user_headers, 1):
            wu.cell(row=3, column=c, value=h)
        style_header_row(wu, 3, len(user_headers))
        wu.freeze_panes = "A4"
        
        for r, u in enumerate(users, start=4):
            alt = (r % 2 == 0)
            row = [r - 3, u.get("username", ""), u.get("full_name", ""), u.get("role", ""),
                   u.get("office", ""), "Yes" if u.get("active", True) else "No", str(u.get("last_login", ""))[:16]]
            for c, val in enumerate(row, 1):
                wu.cell(row=r, column=c, value=val)
            style_data_row(wu, r, len(user_headers), alt)
    
    # Saved Offices sheet
    if "saved_offices" in export_items:
        offices = _export_saved_offices()
        wo = wb.create_sheet("Saved Offices")
        wo.sheet_view.showGridLines = False
        add_title(wo, "Saved Offices", f"Exported {now}")
        
        office_headers = ["#", "Office Slug", "Office Name", "Created By", "Created At"]
        for c, h in enumerate(office_headers, 1):
            wo.cell(row=3, column=c, value=h)
        style_header_row(wo, 3, len(office_headers))
        
        for r, o in enumerate(offices, start=4):
            alt = (r % 2 == 0)
            row = [r - 3, o.get("office_slug",""), o.get("office_name",""), o.get("created_by",""), str(o.get("created_at",""))[:16]]
            for c, val in enumerate(row, 1):
                wo.cell(row=r, column=c, value=val)
            style_data_row(wo, r, len(office_headers), alt)
    
    # Office Traffic sheet
    if "office_traffic" in export_items:
        traffic = _export_office_traffic()
        if traffic:
            wt = wb.create_sheet("Office Traffic")
            wt.sheet_view.showGridLines = False
            add_title(wt, "Office Traffic Log", f"Exported {now}")
            
            traffic_headers = ["#", "Office", "Event Type", "Document ID", "Client Username", "Scanned At"]
            for c, h in enumerate(traffic_headers, 1):
                wt.cell(row=3, column=c, value=h)
            style_header_row(wt, 3, len(traffic_headers))
            
            for r, t in enumerate(traffic, start=4):
                alt = (r % 2 == 0)
                row = [r - 3, t.get("office_name", ""), t.get("event_type", ""), t.get("doc_id", ""),
                       t.get("client_username", ""), str(t.get("scanned_at", ""))[:16]]
                for c, val in enumerate(row, 1):
                    wt.cell(row=r, column=c, value=val)
                style_data_row(wt, r, len(traffic_headers), alt)
    
    # Remove default empty sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def restore_from_excel(excel_bytes: bytes, mode: str = "merge") -> dict:
    """Restore data from an Excel workbook. Returns summary dict."""
    from openpyxl import load_workbook
    from io import BytesIO
    
    wb = load_workbook(BytesIO(excel_bytes), read_only=True, data_only=True)
    
    summary = {
        "mode":          mode,
        "documents":     0,
        "users":         0,
        "routing_slips": 0,
        "saved_offices": 0,
        "office_traffic": 0,
        "skipped":       0,
        "errors":        [],
    }
    
    # Restore Documents
    if "Documents" in wb.sheetnames:
        ws = wb["Documents"]
        docs = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            if row and row[0]:  # Has row number
                doc = {
                    "id": str(row[12]) if row[12] else None,  # Reference No.
                    "date_received": str(row[1]) if row[1] else "",
                    "received_by": str(row[2]) if row[2] else "",
                    "sender_org": str(row[3]) if row[3] else "",
                    "sender_name": str(row[4]) if row[4] else "",
                    "doc_name": str(row[5]) if row[5] else "",
                    "referred_to": str(row[6]) if row[6] else "",
                    "forwarded_to": str(row[7]) if row[7] else "",
                    "date_released": str(row[8]) if row[8] else "",
                    "status": str(row[9]) if row[9] else "Pending",
                    "category": str(row[10]) if row[10] else "",
                    "doc_id": str(row[11]) if row[11] else "",
                    "notes": str(row[12]) if row[12] else "",
                }
                if doc["id"]:
                    docs.append(doc)
        summary["documents"] = _restore_documents(docs, mode, summary)
    
    # Restore Users
    if "Users" in wb.sheetnames:
        ws = wb["Users"]
        users = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            if row and row[0]:  # Has row number
                user = {
                    "username": str(row[1]) if row[1] else None,
                    "full_name": str(row[2]) if row[2] else "",
                    "role": str(row[3]) if row[3] else "staff",
                    "office": str(row[4]) if row[4] else "",
                    "active": str(row[5]).lower() == "yes" if row[5] else True,
                }
                if user["username"]:
                    users.append(user)
        summary["users"] = _restore_users(users, mode, summary)
    
    # Restore Routing Slips
    if "Routing Slips" in wb.sheetnames:
        ws = wb["Routing Slips"]
        slips = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            if row and row[0]:  # Has row number
                slip = {
                    "id": str(row[1]) if row[1] else None,  # Slip No.
                    "slip_no": str(row[1]) if row[1] else "",
                    "slip_date": str(row[2]) if row[2] else "",
                    "destination": str(row[3]) if row[3] else "",
                    "prepared_by": str(row[4]) if row[4] else "",
                    "notes": str(row[5]) if row[5] else "",
                }
                if slip["id"]:
                    slips.append(slip)
        summary["routing_slips"] = _restore_routing_slips(slips, mode, summary)
    
    # Restore Saved Offices
    if "Saved Offices" in wb.sheetnames:
        ws = wb["Saved Offices"]
        offices = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            if row and row[0]:  # Has row number
                office = {
                    "office_slug": str(row[1]) if row[1] else None,
                    "office_name": str(row[2]) if row[2] else "",
                    "created_by": str(row[3]) if row[3] else "",
                }
                if office["office_slug"]:
                    offices.append(office)
        summary["saved_offices"] = _restore_saved_offices(offices, mode, summary)
    
    # Restore Office Traffic
    if "Office Traffic" in wb.sheetnames:
        ws = wb["Office Traffic"]
        traffic = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            if row and row[0]:  # Has row number
                t = {
                    "office_slug": str(row[1]) if row[1] else None,
                    "office_name": str(row[2]) if row[2] else "",
                    "event_type": str(row[3]) if row[3] else "",
                    "doc_id": str(row[4]) if row[4] else "",
                    "client_username": str(row[5]) if row[5] else "",
                    "scanned_at": str(row[6]) if row[6] else "",
                }
                if t["office_slug"]:
                    traffic.append(t)
        summary["office_traffic"] = _restore_office_traffic(traffic, mode, summary)
    
    wb.close()
    return summary


def _export_documents() -> list[dict]:
    """Export ALL documents including soft-deleted ones."""
    if USE_DB:
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute("SELECT data FROM documents ORDER BY created_at DESC")
                    return [row["data"] for row in cur.fetchall()]
        except Exception as e:
            return []
    return load_docs(include_deleted=True)


def _export_users() -> list[dict]:
    """Export users — password hashes included so accounts survive restore."""
    if USE_DB:
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        """SELECT username, password_hash, full_name, role,
                                  active, COALESCE(office,'') AS office,
                                  created_at, last_login
                           FROM users ORDER BY created_at"""
                    )
                    return [dict(r) for r in cur.fetchall()]
        except Exception as e:
            return []
    import os, json as _json
    if os.path.exists("users.json"):
        with open("users.json") as f:
            return _json.load(f)
    return []


def _export_routing_slips() -> list[dict]:
    if USE_DB:
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute("SELECT * FROM routing_slips ORDER BY created_at")
                    rows = cur.fetchall()
                    result = []
                    for r in rows:
                        d = dict(r)
                        # doc_ids is JSONB — ensure it's a list
                        if isinstance(d.get("doc_ids"), str):
                            d["doc_ids"] = json.loads(d["doc_ids"])
                        d["created_at"] = str(d["created_at"]) if d.get("created_at") else ""
                        result.append(d)
                    return result
        except Exception as e:
            return []
    import os
    if os.path.exists("routing_slips.json"):
        with open("routing_slips.json") as f:
            slips = json.load(f)
        return list(slips.values())
    return []


def _export_saved_offices() -> list[dict]:
    if USE_DB:
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT office_slug, office_name, created_by, created_at FROM saved_offices"
                    )
                    rows = [dict(r) for r in cur.fetchall()]
                    for r in rows:
                        r["created_at"] = str(r["created_at"]) if r.get("created_at") else ""
                    return rows
        except Exception as e:
            return []
    import os
    if os.path.exists("saved_offices.json"):
        with open("saved_offices.json") as f:
            offices = json.load(f)
        return [{"office_slug": k, "office_name": v["office_name"],
                 "created_by": v.get("created_by", "")} for k, v in offices.items()]
    return []


def _export_office_traffic() -> list[dict]:
    if USE_DB:
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        """SELECT office_slug, office_name, event_type,
                                  doc_id, client_username, scanned_at
                           FROM office_traffic ORDER BY scanned_at"""
                    )
                    rows = [dict(r) for r in cur.fetchall()]
                    for r in rows:
                        r["scanned_at"] = str(r["scanned_at"])[:19] if r.get("scanned_at") else ""
                    return rows
        except Exception as e:
            return []
    return []


# ── Restore ───────────────────────────────────────────────────────────────────

def restore_backup(backup: dict, mode: str = "merge") -> dict:
    """
    Restore from a backup dict.
    mode='merge'   — keep existing records, only add missing ones (safe)
    mode='replace' — wipe tables and reload everything (full restore)
    Returns a summary dict with counts of what was restored.
    """
    version = backup.get("meta", {}).get("version", "1")
    summary = {
        "mode":          mode,
        "documents":     0,
        "users":         0,
        "routing_slips": 0,
        "saved_offices": 0,
        "office_traffic": 0,
        "skipped":       0,
        "errors":        [],
    }

    if mode == "replace" and USE_DB:
        _wipe_tables()

    summary["documents"]     = _restore_documents(backup.get("documents", []), mode, summary)
    summary["users"]         = _restore_users(backup.get("users", []), mode, summary)
    summary["routing_slips"] = _restore_routing_slips(backup.get("routing_slips", []), mode, summary)
    summary["saved_offices"] = _restore_saved_offices(backup.get("saved_offices", []), mode, summary)
    summary["office_traffic"] = _restore_office_traffic(backup.get("office_traffic", []), mode, summary)

    return summary


def _wipe_tables():
    """Delete all rows from restorable tables — only used in replace mode."""
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM documents")
                cur.execute("DELETE FROM routing_slips")
                cur.execute("DELETE FROM saved_offices")
                cur.execute("DELETE FROM office_traffic")
                # Note: we do NOT wipe users in replace mode for safety
            conn.commit()
    except Exception as e:
        pass


def _restore_documents(docs: list, mode: str, summary: dict) -> int:
    count = 0
    for doc in docs:
        if not isinstance(doc, dict) or not doc.get("id"):
            summary["errors"].append(f"Skipped invalid document: {str(doc)[:60]}")
            summary["skipped"] += 1
            continue
        try:
            existing = get_doc(doc["id"])
            if existing and mode == "merge":
                summary["skipped"] += 1
                continue
            save_doc(doc)
            count += 1
        except Exception as e:
            summary["errors"].append(f"Doc {doc.get('id','?')}: {e}")
    return count


def _restore_users(users: list, mode: str, summary: dict) -> int:
    if not users:
        return 0
    count = 0
    
    if USE_DB:
        for u in users:
            if not u.get("username"):
                continue
            try:
                with get_conn() as conn:
                    with conn.cursor() as cur:
                        if mode == "merge":
                            cur.execute("SELECT 1 FROM users WHERE username=%s", (u["username"],))
                            if cur.fetchone():
                                summary["skipped"] += 1
                                continue
                        cur.execute(
                            """INSERT INTO users
                                   (username, password_hash, full_name, role, active, office)
                               VALUES (%s,%s,%s,%s,%s,%s)
                               ON CONFLICT (username) DO UPDATE SET
                                   full_name = EXCLUDED.full_name,
                                   role      = EXCLUDED.role,
                                   office    = EXCLUDED.office""",
                            (
                                u["username"],
                                u.get("password_hash", ""),
                                u.get("full_name", ""),
                                u.get("role", "staff"),
                                u.get("active", True),
                                u.get("office", ""),
                            )
                        )
                    conn.commit()
                count += 1
            except Exception as e:
                summary["errors"].append(f"User {u.get('username','?')}: {e}")
    else:
        # JSON file mode
        import os
        existing = {}
        if os.path.exists("users.json"):
            with open("users.json") as f:
                existing = json.load(f)
        
        for u in users:
            if not u.get("username"):
                continue
            username = u["username"]
            if mode == "merge" and username in existing:
                summary["skipped"] += 1
                continue
            existing[username] = u
            count += 1
        
        with open("users.json", "w") as f:
            json.dump(existing, f, indent=2, default=str)
    
    return count


def _restore_routing_slips(slips: list, mode: str, summary: dict) -> int:
    if not slips:
        return 0
    count = 0
    
    if USE_DB:
        for slip in slips:
            if not slip.get("id"):
                continue
            try:
                with get_conn() as conn:
                    with conn.cursor() as cur:
                        if mode == "merge":
                            cur.execute("SELECT 1 FROM routing_slips WHERE id=%s", (slip["id"],))
                            if cur.fetchone():
                                summary["skipped"] += 1
                                continue
                        cur.execute(
                            """INSERT INTO routing_slips
                                   (id, slip_no, destination, prepared_by,
                                    doc_ids, notes, slip_date, time_from, time_to)
                               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
                               ON CONFLICT (id) DO NOTHING""",
                            (
                                slip["id"], slip.get("slip_no", ""),
                                slip.get("destination", ""), slip.get("prepared_by", ""),
                                json.dumps(slip.get("doc_ids", [])),
                                slip.get("notes", ""), slip.get("slip_date", ""),
                                slip.get("time_from", ""), slip.get("time_to", ""),
                            )
                        )
                    conn.commit()
                count += 1
            except Exception as e:
                summary["errors"].append(f"Slip {slip.get('id','?')}: {e}")
    else:
        # JSON file mode
        import os
        existing = {}
        if os.path.exists("routing_slips.json"):
            with open("routing_slips.json") as f:
                existing = json.load(f)
        
        for slip in slips:
            if not slip.get("id"):
                continue
            slip_id = slip["id"]
            if mode == "merge" and slip_id in existing:
                summary["skipped"] += 1
                continue
            existing[slip_id] = slip
            count += 1
        
        with open("routing_slips.json", "w") as f:
            json.dump(existing, f, indent=2, default=str)
    
    return count


def _restore_saved_offices(offices: list, mode: str, summary: dict) -> int:
    if not offices:
        return 0
    count = 0
    
    if USE_DB:
        for office in offices:
            if not office.get("office_slug"):
                continue
            try:
                with get_conn() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            """INSERT INTO saved_offices (office_slug, office_name, created_by)
                               VALUES (%s,%s,%s)
                               ON CONFLICT (office_slug) DO NOTHING""",
                            (office["office_slug"], office.get("office_name", ""),
                             office.get("created_by", ""))
                        )
                    conn.commit()
                count += 1
            except Exception as e:
                summary["errors"].append(f"Office {office.get('office_slug','?')}: {e}")
    else:
        # JSON file mode
        import os
        existing = {}
        if os.path.exists("saved_offices.json"):
            with open("saved_offices.json") as f:
                existing = json.load(f)
        
        for office in offices:
            if not office.get("office_slug"):
                continue
            slug = office["office_slug"]
            if mode == "merge" and slug in existing:
                summary["skipped"] += 1
                continue
            existing[slug] = {
                "office_name": office.get("office_name", ""),
                "created_by": office.get("created_by", "")
            }
            count += 1
        
        with open("saved_offices.json", "w") as f:
            json.dump(existing, f, indent=2, default=str)
    
    return count


def _restore_office_traffic(traffic: list, mode: str, summary: dict) -> int:
    if not traffic:
        return 0
    if not USE_DB:
        # Office traffic only works with database, skip in JSON file mode
        summary["errors"].append("Office traffic requires database — skipped in local mode")
        return 0
    
    count = 0
    for t in traffic:
        if not t.get("office_slug") or not t.get("scanned_at"):
            continue
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        """INSERT INTO office_traffic
                               (office_slug, office_name, event_type,
                                doc_id, client_username, scanned_at)
                           VALUES (%s,%s,%s,%s,%s,%s)
                           ON CONFLICT DO NOTHING""",
                        (
                            t["office_slug"],
                            t.get("office_name", ""),
                            t.get("event_type", ""),
                            t.get("doc_id", ""),
                            t.get("client_username", ""),
                            t.get("scanned_at", ""),
                        )
                    )
                conn.commit()
            count += 1
        except Exception as e:
            summary["errors"].append(f"Traffic {t.get('office_slug','?')}: {e}")
    return count
